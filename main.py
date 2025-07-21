# IMPORTA BIBLIOTECAS
from datetime import date as dt
from openpyxl import load_workbook
from datetime import datetime
import time as t
import shutil

wb = load_workbook('dados.xlsx') #Carregar arquivo excell TROQUE PELO NOME DO ARQUIVO QUE VOCE VAI ABRIR
ws = wb.active #Ativar o excell
linha = ws.max_row #Ler o maximo de linhas

while True: #Estrutura repeticao
    print("\nDeixe em branco para fechar o programa!")  # entradas de valores
    servico = input("Servico: ").upper()
    if servico == '':break

    while True:
        valor = input("Valor: ").replace(",", ".").strip()
        try:
            valor = float(valor)
            if valor == float(valor):break
        except ValueError: print("Erro digite novamento")
    pagamento = input("Tipo de pagamento: ").upper()
    while pagamento == '':
        print("valor nao nulo identificado")
        pagamento = input("Tipo de pagamento: ").upper()
    observacoes = input("Observacoes: ").upper()
    agora = datetime.now() 
    hora_atual = agora.strftime("%H:%M")

    while all(cell.value is None for cell in ws[linha]):linha-=1
    linha+=1

    hoje1 = dt.today()
    dataHoje = hoje1.strftime("%d/%m/%Y")

    ws.cell(row=linha, column=1).value = dataHoje
    ws.cell(row=linha, column=2).value = servico
    ws.cell(row=linha, column=3).value = valor
    ws.cell(row=linha, column=4).value = pagamento
    ws.cell(row=linha,  column=5).value = observacoes
    ws.cell(row=linha, column=6).value = hora_atual
    wb.save('dados.xlsx') # BOTE O NOME QUE DESEJA SALVAR

print("\nObrigado por usar!\n")
t.sleep(1)

from salvar import Backup
Backup()


